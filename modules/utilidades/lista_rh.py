from openpyxl import load_workbook

class Lista_rh:
    __arquivo = "listaRh.xlsx"
    def __init__(self):
        self.__funcionarios = []
        self.carregar_funcionarios()

    def carregar_funcionarios(self):
        wb = load_workbook(self.__arquivo)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            local = None

            if row != None:
                if row[4] is not None:
                    local = row[4]

            registro = {
                'local_de_trabalho': local,
                'matricula': row[0],
                'nome': row[2],
                'cargo': row[3]
            }

            self.__funcionarios.append(registro)
    
    def buscar_por_matricula(self, matricula):
        matricula_temp = matricula.split('-')[0]
        matricula_temp_sem_ultimo = matricula_temp[:-1]

        resultado1 = next((funcionario for funcionario in self.__funcionarios if str(funcionario['matricula']) == matricula_temp_sem_ultimo), None)
        resultado2 = next((funcionario for funcionario in self.__funcionarios if str(funcionario['matricula']) == matricula_temp), None)

        if resultado1:
            return resultado1
        
        if resultado2:
            return resultado2
